import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table
from openpyxl.utils import column_index_from_string
from typing import Optional, Union
from pathlib import Path


def append_dataframe_to_excel(
    dataframe: pd.DataFrame, excel_file: Union[str, Path], sheet_name: str
) -> None:
    """
    Appends a pandas DataFrame to an existing or new Excel file in a specified sheet.

    Args:
        dataframe (pd.DataFrame): The DataFrame to be appended to the Excel file.
        excel_file (Union[str, Path]): The path to the Excel file where the DataFrame will be appended.
        sheet_name (str): The name of the sheet where the DataFrame should be appended.

    Returns:
        None
    """
    file_path = Path(excel_file)

    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(dataframe.columns))
        for index, row in dataframe.iterrows():
            row_list = list(row)
            ws.append(row_list)
        wb.save(excel_file)
        return

    else:
        wb = load_workbook(file_path)
        if sheet_name in wb:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(list(dataframe.columns))

    for index, row in dataframe.iterrows():
        row_list = list(row)
        ws.append(row_list)

    wb.save(excel_file)
